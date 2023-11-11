# 予めインストール
# pip3 install requests

# ライブラリをインポート
# Excelを扱えるライブラリ
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fonts import Font
from openpyxl.styles.borders import Border, Side
# Webスクレイピング用ライブラリ
import requests
# 全角→半角の正規化処理
import unicodedata
# 生年月日
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta

# 曜日一覧
jpn_weekday = ['月', '火', '水', '木', '金', '土', '日']

# 年末年始休暇(12/29-1/3を前提)
newyear_holidays = [[12, 29], [1, 3]]

# 標準入力
str = input('\n年月を6桁で入力してください (例:2023年11月の場合「202311」と入力)\n→')
# 202311

# 全角→半角の正規化処理
import unicodedata
normalize = unicodedata.normalize("NFKC", str)    # ２０２３１１ >> 202311

# 月初日
first_date = datetime.strptime(normalize + "01", '%Y%m%d') # 20231101 >> 2023-11-01 00:00:00
# 月末日
last_date = first_date + relativedelta(months=1) - timedelta(days=1)

# Excelの操作
wb = openpyxl.Workbook()
sht = wb.active

# 背景色
bgcolor_red = PatternFill(patternType='solid', fgColor='FF66FF', bgColor='FF66FF')
bgcolor_blue = PatternFill(patternType='solid', fgColor='00B0F0', bgColor='00B0F0')
bgcolor_none = PatternFill(fill_type=None)

# 罫線
side = Side(style='thin', color='000000')
border = Border(top=side, bottom=side, left=side, right=side)

# タイトル
sht.cell(1, 1).value = f'{first_date.year}年{first_date.month}月 予定表'
sht.cell(1, 1).font = Font(size=22)
# 選択範囲で中央
for col_cnt in range(1, 12):
    sht.cell(1, col_cnt).alignment = Alignment(horizontal="centerContinuous")

sht.column_dimensions['A'].width = 6 # A列の幅
sht.column_dimensions['B'].width = 6 # B列の幅

for time_cnt in range(9, 18):
    rng = sht.cell(3, time_cnt - 6)
    rng.value = f'{time_cnt}:00～'
    # 罫線
    rng.border = border
    # 列幅
    sht.column_dimensions[chr(96 + time_cnt - 6)].width = 15

for date_cnt in range(first_date.day, last_date.day + 1):
    buf_date = datetime.strptime(f'{normalize}{date_cnt:02}', '%Y%m%d')
    
    sht.cell(date_cnt + 3, 1).value = date_cnt
    sht.cell(date_cnt + 3, 2).value = jpn_weekday[buf_date.weekday()]
    
    if buf_date.month == newyear_holidays[0][0] and buf_date.day >= newyear_holidays[0][1]:
        # 年末
        buf_bgcolor = bgcolor_red

    elif buf_date.month == newyear_holidays[1][0] and buf_date.day <= newyear_holidays[1][1]:
        # 年始
        buf_bgcolor = bgcolor_red

    elif buf_date.weekday() == 5: #土曜の場合
        # webから祝日の情報を取得する(祝日のみhollidayになる)
        res = requests.get(f'https://s-proj.com/utils/checkHoliday.php?kind=ph&date={normalize}{date_cnt:02}')

        if res.content.decode("utf-8") == "holiday": #祝日の場合
            buf_bgcolor = bgcolor_red
        else:
            buf_bgcolor = bgcolor_blue

    else:
        # webから祝日の情報を取得する
        res = requests.get(f'https://s-proj.com/utils/checkHoliday.php?date={normalize}{date_cnt:02}')
        
        if res.content.decode("utf-8") == "holiday": #日曜・祝日の場合
            buf_bgcolor = bgcolor_red
        else: # 平日
            buf_bgcolor = bgcolor_none
            sht.row_dimensions[date_cnt + 3].height = 50
        
    for col_cnt in range(1, 11 + 1):
        rng = sht.cell(date_cnt + 3, col_cnt)
        # 背景色を設定する
        rng.fill = buf_bgcolor
        # 中央揃え
        rng.alignment = Alignment(horizontal="center", vertical="center")
        # 罫線
        rng.border = border

# 印刷の設定
sht.print_area = "A1:K34"
# 余白
sht.page_margins.left = 0.2
sht.page_margins.right = 0.2
sht.page_margins.top = 0.2
sht.page_margins.bottom = 0.2
sht.page_margins.header = 0
sht.page_margins.footer = 0
sht.page_setup.fitToPage = True
# 横をすべて1ページにおさめて印刷
sht.page_setup.fitToWidth = 1
# 縦をすべて1ページにおさめて印刷
sht.page_setup.fitToHeight = 1
# 横方向中央に寄せて印刷
sht.print_options.horizontalCentered = True

save_filename = f'月次予定表_{normalize}.xlsx'
wb.save(save_filename)

print("月次予定表のExcelデータが正常に作成されました。")
print(f'ファイル名:{save_filename}')
import os
os.system('PAUSE')
